from openpyxl import Workbook
wb = Workbook()
sheet = wb.active
ws = wb.create_sheet("Student Course Information")

row = 1
ws.cell(row, 1, 'Roll Number')
ws.cell(row, 2, 'Name')
ws.cell(row, 3, 'Semester')
ws.cell(row, 4, 'Course 1')
ws.cell(row, 5, 'Course 2')
ws.cell(row, 6, 'Course 3')
more = True
while(more):
  print('\n')
  ++row
  roll = input('Enter your Roll Number : ')
  ws.cell(row, 1, roll)
  name = input('Enter your Name : ')
  ws.cell(row, 2, name)
  sem = input('Enter your semester out of 2, 4, 6 or 8 : ')
  ws.cell(row, 3, sem)
  year = int(sem)//2
  branch = input('Choose your branch CSE or ECE : ')
  if year == 1:
    if branch == 'CSE':
      print('You have following choices, choose 3 out of them :')
      print('CS101 : C programming\nCS102 : Java programming\nCS103 : Python programming\nCS104 : Data Structures')
      print('Note that CS104 is compulsary\nChoose y or n for the courses')
      c1 = input('CS101 : C programming : ')
      if c1 == 'n':
        ws.cell(row, 4, 'CS102 : Java programming')
        ws.cell(row, 5, 'CS103 : Python programming')
        ws.cell(row, 6, 'CS104 : Data Structures')
      else:
        c2 = input('CS102 : Java programming : ')
        if c2 == 'n':
          ws.cell(row, 4, 'CS101 : C programming')
          ws.cell(row, 5, 'CS103 : Python programming')
          ws.cell(row, 6, 'CS104 : Data Structures')
        else:
          c3 = input('CS103 : Python programming : ')
          if c3 == 'n':
            ws.cell(row, 4, 'CS101 : C programming')
            ws.cell(row, 5, 'CS102 : Java programming')
            ws.cell(row, 6, 'CS104 : Data Structures')
    else :
      print('You have following choices, choose 2 out of them :')
      print('EC101 : Analog Circuits\nEC102 : Semiconductors\nEC103 : Signals and System\nEC104 : Digital Electronics')
      print('Note that EC103 and EC104 are compulsary\nChoose y or n for the courses')
      c1 = input('EC101 : Analog Circuits : ')
      if c1 == 'n':
        ws.cell(row, 4, 'EC102 : Semiconductors')
        ws.cell(row, 5, 'EC103 : Signals and System')
        ws.cell(row, 6, 'EC104 : Digital Electronics')
      else:
        c2 = input('EC102 : Semiconductors : ')
        if c2 == 'n':
          ws.cell(row, 4, 'EC101 : Analog Circuits')
          ws.cell(row, 5, 'EC103 : Signals and System')
          ws.cell(row, 6, 'EC104 : Digital Electronics')
  elif year == 2:
    if branch == 'CSE':
      print('You have following choices, choose 3 out of them :')
      print('CS201 : Advanced C++\nCS202 : IT Workshop 1\nCS203 : Database Management System\nCS204 : Algorithms')
      print('Note that CS203 and CS204 are compulsary\nChoose y or n for the courses')
      c1 = input('CS201 : Advanced C++ : ')
      if c1 == 'n':
        ws.cell(row, 4, 'CS202 : IT Workshop 1')
        ws.cell(row, 5, 'CS203 : Database Management System')
        ws.cell(row, 6, 'CS204 : Algorithms')
      else:
        c2 = input('CS202 : IT Workshop 1 : ')
        if c2 == 'n':
          ws.cell(row, 4, 'CS201 : Advanced C++')
          ws.cell(row, 5, 'CS203 : Database Management System')
          ws.cell(row, 6, 'CS204 : Algorithms')
    else:
      print('You have following choices, choose 3 out of them :')
      print('EC201 : Digital Communications\nEC202 : Electromagnetics\nEC203 : Analog Circuits\nEC204 : Analog Circuits Lab')
      print('Note that EC203 and EC204 are compulsary\nChoose y or n for the courses')
      c1 = input('EC201 : Digital Communications : ')
      if c1 == 'n':
        ws.cell(row, 4, 'EC202 : Electromagnetics')
        ws.cell(row, 5, 'EC203 : Analog Circuits')
        ws.cell(row, 6, 'EC204 : Analog Circuits Lab')
      else:
        c2 = input('EC202 : Electromagnetics : ')
        if c2 == 'n':
          ws.cell(row, 4, 'EC201 : Digital Communications')
          ws.cell(row, 5, 'EC203 : Analog Circuits')
          ws.cell(row, 6, 'EC204 : Analog Circuits Lab')
  elif year == 3:
    if branch == 'CSE':
      print('You have 3 courses, all of them are compulsary :')
      print('CS301 : IT Workshop 1\nCS302 : Database Management System\nCS303 : Algorithms')
      ws.cell(row, 4, 'CS301 : IT Workshop 1')
      ws.cell(row, 5, 'CS302 : Database Management System')
      ws.cell(row, 6, 'CS303 : Algorithms')
    else:
      print('You have 3 courses, all of them are compulsary :')
      print('EC301 : Electromagnetics\nEC302 : Analog Circuits\nEC303 : Analog Circuits Lab')
      ws.cell(row, 4, 'EC301 : Electromagnetics')
      ws.cell(row, 5, 'EC302 : Analog Circuits')
      ws.cell(row, 6, 'EC303 : Analog Circuits Lab')
  elif year == 4:
    if branch == 'CSE':
      print('You have following choices, choose 3 out of them :')
      print('CS401 : Machine Learning\nCS402 : Cloud Computing\nCS403 : Quantum Computing\nCS404 : Distributed Systems')
      c1 = input('CS401 : Machine Learning : ')
      if c1 == 'n':
        ws.cell(row, 4, 'CS402 : Cloud Computing')
        ws.cell(row, 5, 'CS403 : Quantum Computing')
        ws.cell(row, 6, 'CS404 : Distributed Systems')
      else:
        c2 = input('CS402 : Cloud Computing : ')
        if c2 == 'n':
          ws.cell(row, 4, 'CS401 : Machine Learning')
          ws.cell(row, 5, 'CS403 : Quantum Computing')
          ws.cell(row, 6, 'CS404 : Distributed Systems')
        else:
          c3 = input('CS403 : Quantum Computing : ')
          if c3 == 'n':
            ws.cell(row, 4, 'CS401 : Machine Learning')
            ws.cell(row, 5, 'CS402 : Cloud Computing')
            ws.cell(row, 6, 'CS404 : Distributed Systems')
          else:
            c3 = input('CS404 : Distributed Systems : ')
            if c4 == 'n':
              ws.cell(row, 4, 'CS401 : Machine Learning')
              ws.cell(row, 5, 'CS402 : Cloud Computing')
              ws.cell(row, 6, 'CS403 : Quantum Computing')
    else:
      print('You have following choices, choose 3 out of them :')
      print('EC401 : Probability and Stochastic Process\nEC402 : Linear Algebra and Optimization Theory\nEC403 : Advanced Digital Signal Processing\nEC404 : Advanced Digital Signal Processing Lab')
      print('EC403 and EC404 are compulsary')
      c1 = input('EC401 : Probability and Stochastic Process : ')
      if c1 == 'n':
        ws.cell(row, 4, 'EC402 : Linear Algebra and Optimization Theory')
        ws.cell(row, 5, 'EC403 : Advanced Digital Signal Processing')
        ws.cell(row, 6, 'EC404 : Advanced Digital Signal Processing Lab')
      else:
        c2 = input('EC402 : Linear Algebra and Optimization Theory : ')
        if c2 == 'n':
          ws.cell(row, 4, 'EC401 : Probability and Stochastic Process')
          ws.cell(row, 5, 'EC403 : Advanced Digital Signal Processing')
          ws.cell(row, 6, 'EC404 : Advanced Digital Signal Processing Lab')
  x = input('Do you want to make more entries? y/n : ')
  if x == 'n':
    more = False
print('\n')
type(ws.rows)
for curr_row in ws.rows:
  for cell in curr_row:
    print(cell.value, end = '\t\t')
  print('\n')
